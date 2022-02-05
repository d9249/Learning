from email.mime import base
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from PIL import Image as IMG
from .models import Category,Crack,CrackObj

def facility(wb,pk):
  category = Category.objects.get(pk=pk)
  
  grayFill = PatternFill(start_color='CCCCCC',
                        end_color='CCCCCC', fill_type='solid')
  sheet = wb.worksheets[0]
  sheet.title = '시설물 현황'

  sheet = wb['시설물 현황']
  sheet['B2'] = "□ 시설물 현황"
  sheet['B3'] = "가. 일반현황"
  sheet['B12'] = "나. 전경사진"

  file_path = category.frontView.url
  path = file_path[1:]
  img = Image(path)
  sheet.add_image(img, "B13")

  sheet['B4'] = '시설물명'
  sheet['B5'] = '시설물위치'
  sheet['B6'] = '용도'
  sheet['B7'] = '구조형식'

  sheet['B4'].fill = grayFill
  sheet['B5'].fill = grayFill
  sheet['B6'].fill = grayFill
  sheet['B7'].fill = grayFill

  
  sheet.merge_cells('C4:D4')
  sheet.merge_cells('C5:D5')
  sheet.merge_cells('C6:D6')
  sheet.merge_cells('C7:D7')
  sheet['C4'] = category.facilityName
  sheet['C5'] = category.facilityNo
  sheet['C6'] = category.usage
  sheet['C7'] = category.structuralForm

  sheet['E4'] = "시설물번호"
  sheet['E5'] = "준공일자"
  sheet['E6'] = "시설물규모"
  sheet['E7'] = "부대시설"

  sheet['E4'].fill = grayFill
  sheet['E5'].fill = grayFill
  sheet['E6'].fill = grayFill
  sheet['E7'].fill = grayFill

  sheet.merge_cells('F4:G4')
  sheet.merge_cells('F5:G5')
  sheet.merge_cells('F6:G6')
  sheet.merge_cells('F7:G7')

  sheet['F4'] = category.facilityNo
  sheet['F5'] = category.completionDate
  sheet['F6'] = category.facilityStructure
  sheet['F7'] = category.amenities

  sheet['B8'] = '종별'
  sheet['D8'] = '전차안전등급'
  sheet['F8'] = '점검결과안전등급'

  sheet['B8'].fill = grayFill
  sheet['D8'].fill = grayFill
  sheet['F8'].fill = grayFill

  sheet['C8'] = category.floors
  sheet['E8'] = category.grade
  sheet['G8'] = category.testResults

  sheet.merge_cells('B9:G9')
  sheet.merge_cells('B10:G10')

  sheet['B9'] = '규모 및 제원 추가사항'
  sheet['B9'].fill = grayFill

  sheet['B10'] = category.plus

  sheet['B12'] = '나. 전경사진'
  path = category.frontView.url
  path = path[1:]
  img = Image(path)
  sheet.add_image(img,"B13")
  sheet.sheet_view.view = "pageBreakPreview"
  for row in sheet.rows:
      for cell in row:
          cell.alignment = Alignment(horizontal="center", vertical="center")
  return wb


def looks(wb,pk):
  baseWidth = 210
  imgCell = 2
  infoCell = 10
  sheet = wb.create_sheet("외관조사사진", 1)
  sheet = wb['외관조사사진']
  category = Category.objects.get(pk=pk)
  cracks = Crack.objects.filter(category__facilityName__icontains=category.facilityName)
  sheet.column_dimensions["A"].width = 1
  sheet.column_dimensions["D"].width = 1
  
  for crack in cracks:
    crackObj = CrackObj.objects.filter(parent=crack.id)
    print(crackObj)
    numbering = crackObj.count()
    if numbering < 3:
      numbering = 0
    else:
      numbering = numbering-2
    crackObj = crackObj[numbering:]
    print(numbering)
    print(crackObj)
    
    for crackObj in crackObj:
      path = crackObj.image.url[1:]
      img = IMG.open(path) # 사진의 비율을 알기 위한 변수 PIL 라이브러리
      wpercent = baseWidth/float(img.size[0])
      hsize = int((float(img.size[1])* float(wpercent)))

      flatPath = crackObj.flatting_image.url[1:]
      flatImg = IMG.open(flatPath) # 사진의 비율을 알기 위한 변수 PIL 라이브러리
      flatwPercent = baseWidth/float(img.size[0])
      flathSize = int((float(flatImg.size[1])* float(flatwPercent)))
      if flathSize > 160:
        flathSize= 160

      image = openpyxl.drawing.image.Image(path) # 엑셀에 이미지 삽입을 위한 변수 openpyxl 라이브러리
      flatImage = openpyxl.drawing.image.Image(flatPath)
      
      image.width = baseWidth
      image.height = hsize
      
      flatImage.width = baseWidth
      flatImage.height = flathSize
 
      sheet.add_image(image,'B' + str(imgCell))
      sheet.add_image(flatImage, 'C'+ str(imgCell))

      sheet.column_dimensions["B"].width = 27
      sheet.column_dimensions["C"].width = 27

      sheet['B'+str(infoCell)] = '사진번호: ' + str(crackObj.id)
      sheet['B'+str(infoCell+1)] = '위치: ' + str(crack.floor) + str(crack.location)
      sheet['B'+str(infoCell+2)] = '점검내용: ' + str(crack.desc)
      sheet['C'+str(infoCell+2)] = '손상규모: ' + str(crackObj.crackLength)
      sheet['B'+str(infoCell+3)] = '발생원인: ' + str(crack.cause)
      sheet['C'+str(infoCell+3)] = '진행유무: ' + str(crack.progress)
      imgCell += 11
      infoCell += 11
      sheet.sheet_view.view = "pageBreakPreview"
  return wb
