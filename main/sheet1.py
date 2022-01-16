import openpyxl
from openpyxl.styles import Alignment, PatternFill
from .models import Category

def sheet(wb,pk):
  category = Category.objects.get(pk=pk)
  grayFill = PatternFill(start_color='CCCCCC',
                        end_color='CCCCCC', fill_type='solid')
  ws1 = wb.create_sheet("시설물 현황", 0)


  ws1 = wb.active
  ws1['B2'] = "□ 시설물 현황"
  ws1['B3'] = "가. 일반현황"

  ws1['B4'] = '시설물명'
  ws1['B5'] = '시설물위치'
  ws1['B6'] = '용도'
  ws1['B7'] = '구조형식'

  ws1['B4'].fill = grayFill
  ws1['B5'].fill = grayFill
  ws1['B6'].fill = grayFill
  ws1['B7'].fill = grayFill

  
  ws1.merge_cells('C4:D4')
  ws1.merge_cells('C5:D5')
  ws1.merge_cells('C6:D6')
  ws1.merge_cells('C7:D7')
  ws1['C4'] = category.facilityName
  ws1['C5'] = category.facilityNo
  ws1['C6'] = category.usage
  ws1['C7'] = category.structuralForm

  ws1['E4'] = "시설물번호"
  ws1['E5'] = "준공일자"
  ws1['E6'] = "시설물규모"
  ws1['E7'] = "부대시설"

  ws1['E4'].fill = grayFill
  ws1['E5'].fill = grayFill
  ws1['E6'].fill = grayFill
  ws1['E7'].fill = grayFill

  ws1.merge_cells('F4:G4')
  ws1.merge_cells('F5:G5')
  ws1.merge_cells('F6:G6')
  ws1.merge_cells('F7:G7')

  ws1['F4'] = category.facilityNo
  ws1['F5'] = category.completionDate
  ws1['F6'] = category.facilityStructure
  ws1['F7'] = category.amenities

  ws1['B8'] = '종별'
  ws1['D8'] = '전차안전등급'
  ws1['F8'] = '점검결과안전등급'

  ws1['B8'].fill = grayFill
  ws1['D8'].fill = grayFill
  ws1['F8'].fill = grayFill

  ws1['C8'] = category.floors
  ws1['E8'] = category.grade
  ws1['G8'] = category.testResults

  ws1.merge_cells('B9:G9')
  ws1.merge_cells('B10:G10')

  ws1['B9'] = '규모 및 제원 추가사항'
  ws1['B9'].fill = grayFill

  ws1['B10'] = category.plus

  for row in ws1.rows:
      for cell in row:
          cell.alignment = Alignment(horizontal="center", vertical="center")
  return wb